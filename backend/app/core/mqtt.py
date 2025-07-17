# app/core/mqtt.py

import json
import ssl
import logging
import os
from datetime import datetime

# ---  CÁC THƯ VIỆN GHI FILE EXCEL ---
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# ------------------------------------

# --- THÊM CÁC THƯ VIỆN GOOGLE SHEETS API MỚI ---
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
# ----------------------------------------------

logger = logging.getLogger(__name__)
mqtt_client = None

# ✅ Tạo đường dẫn tuyệt đối đến thư mục logs (giữ lại nếu bạn muốn ghi log khác)
BASE_DIR = os.path.dirname(os.path.dirname(__file__))  # trỏ tới app/
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

# ✅ Đường dẫn tới file Excel log
EXCEL_FILE = os.path.join(LOG_DIR, "mqtt_data.xlsx")

# ✅ Đường dẫn tới file credentials.json
# Đặt file credentials.json trong thư mục app/core/ (cùng cấp với mqtt.py)
CREDENTIALS_FILE = os.path.join(os.path.dirname(__file__), "credentials.json")

# ✅ ID của Google Sheet của bạn - THAY THẾ BẰNG ID THỰC CỦA BẠN!
# Bạn có thể tìm thấy ID này trong URL của Google Sheet:
# https://docs.google.com/spreadsheets/d/YOUR_SPREADSHEET_ID_HERE/edit
SPREADSHEET_ID = "1_hfhka2P0cM3w5q80MDqNIQ1ut2RzXMJ-T1DNEYzOCc" # <<< THAY THẾ CHỖ NÀY !!!

def get_mqtt_client():
    global mqtt_client
    if mqtt_client is None:
        from paho.mqtt.client import Client
        mqtt_client = Client()

        mqtt_client.on_connect = on_connect
        mqtt_client.on_message = on_message

        # Đường dẫn đến ca.pem (đảm bảo file này vẫn đúng vị trí)
        cert_path = os.path.join(os.path.dirname(__file__), "certs", "emqxsl_ca.pem")

        # Thiết lập TLS chỉ dùng CA
        mqtt_client.tls_set(
            ca_certs=cert_path,
            cert_reqs=ssl.CERT_REQUIRED,
            tls_version=ssl.PROTOCOL_TLSv1_2,
        )
        mqtt_client.tls_insecure_set(False)

        # Kết nối đến broker
        mqtt_client.connect("scalemodelvn.com", port=8883)
        mqtt_client.loop_start()

    return mqtt_client

def on_connect(client, userdata, flags, rc):
    if rc == 0:
        logger.info("✅ MQTT connected securely with CA cert only")
        client.subscribe("secure/topic")  # Thay topic nếu cần
    else:
        logger.error(f"❌ MQTT connection failed with code {rc}")

def on_message(client, userdata, msg):
    payload = msg.payload.decode()
    topic = msg.topic
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    logger.info(f"📩 MQTT Message: {topic} -> {payload}")
    # Gọi hàm mới để ghi vào Google Sheet
    log_to_google_sheet(timestamp, topic, payload)
    log_to_excel(timestamp, topic, payload) # Gọi lại hàm ghi Excel

def log_to_excel(timestamp: str, topic: str, payload: str):
    try:
        # Tạo mới file nếu chưa tồn tại
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "MQTT Logs"
            ws.append(["Timestamp", "Topic", "Payload"])
            logger.info(f"✅ Đã tạo file Excel mới: {EXCEL_FILE}")
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        # Thêm dòng dữ liệu mới
        ws.append([timestamp, topic, payload])

        # Auto-resize cột (tùy chọn, có thể bỏ qua nếu không cần)
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(EXCEL_FILE)
        logger.info(f"✅ Đã ghi dữ liệu vào file Excel cục bộ: {EXCEL_FILE}")
    except Exception as e:
        logger.error(f"❌ Lỗi khi ghi vào file Excel: {e}")

def log_to_google_sheet(timestamp: str, topic: str, payload: str):
    # Phạm vi quyền mà ứng dụng của bạn cần
    # Chỉ cần quyền ghi vào Google Sheets
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    creds = None
    try:
        # Tải thông tin xác thực từ file JSON
        creds = service_account.Credentials.from_service_account_file(
            CREDENTIALS_FILE, scopes=SCOPES)
    except FileNotFoundError:
        logger.error(f"❌ Lỗi: Không tìm thấy file thông tin xác thực tại: {CREDENTIALS_FILE}. Vui lòng kiểm tra đường dẫn và tên file.")
        return
    except Exception as e:
        logger.error(f"❌ Lỗi khi tải thông tin xác thực: {e}")
        return

    try:
        # Xây dựng đối tượng dịch vụ Google Sheets API
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        # Dữ liệu bạn muốn ghi vào Google Sheet
        # Mỗi phần tử trong list con là một ô trong hàng
        values = [[timestamp, topic, payload]]
        body = {
            'values': values
        }

        # Thực hiện việc ghi dữ liệu vào Google Sheet
        # range="Sheet1!A1": Chỉ định tên sheet (mặc định là Sheet1) và ô bắt đầu (A1)
        # valueInputOption="RAW": Dữ liệu được chèn nguyên văn, không qua phân tích cú pháp
        # insertDataOption="INSERT_ROWS": Chèn dữ liệu dưới dạng hàng mới vào cuối sheet
        result = sheet.values().append(
            spreadsheetId=SPREADSHEET_ID,
            range="Sheet1!A1", # Đảm bảo tên sheet này khớp với tên tab trong Google Sheet của bạn
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body=body).execute()

        # Log thông báo thành công
        logger.info(f"✅ Đã ghi dữ liệu vào Google Sheet: {result.get('updates').get('updatedCells')} ô đã được cập nhật.")

    except HttpError as err:
        # Xử lý các lỗi từ Google Sheets API (ví dụ: lỗi quyền truy cập, sheet không tồn tại)
        logger.error(f"❌ Lỗi Google Sheets API: {err}")
    except Exception as e:
        # Xử lý các lỗi không mong muốn khác
        logger.error(f"❌ Đã xảy ra lỗi không mong muốn khi ghi vào Google Sheet: {e}")


def publish(topic: str, payload: dict):
    client = get_mqtt_client()
    client.publish(topic, json.dumps(payload))